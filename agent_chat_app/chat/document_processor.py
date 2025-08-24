import os
import logging
from typing import List, Dict, Any
from pathlib import Path

try:
    import PyPDF2
    from docx import Document as DocxDocument
    import openpyxl
    from langchain.text_splitter import RecursiveCharacterTextSplitter
except ImportError as e:
    logging.warning(f"Some document processing dependencies not available: {e}")

logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Process various document types for RAG"""
    
    def __init__(self, chunk_size: int = 1000, chunk_overlap: int = 200):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            length_function=len,
        )
    
    def extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text
        except Exception as e:
            logger.error(f"Error extracting text from PDF {file_path}: {e}")
            return ""
    
    def extract_text_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        try:
            doc = DocxDocument(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting text from DOCX {file_path}: {e}")
            return ""
    
    def extract_text_from_xlsx(self, file_path: str) -> str:
        """Extract text from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting text from Excel {file_path}: {e}")
            return ""
    
    def extract_text_from_txt(self, file_path: str) -> str:
        """Extract text from TXT file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except Exception as e:
            logger.error(f"Error reading text file {file_path}: {e}")
            return ""
    
    def process_document(self, file_path: str, filename: str = None) -> List[Dict[str, Any]]:
        """Process a document and return chunks with metadata"""
        if not filename:
            filename = os.path.basename(file_path)
        
        file_extension = Path(file_path).suffix.lower()
        
        # Extract text based on file type
        if file_extension == '.pdf':
            text = self.extract_text_from_pdf(file_path)
        elif file_extension == '.docx':
            text = self.extract_text_from_docx(file_path)
        elif file_extension == '.xlsx':
            text = self.extract_text_from_xlsx(file_path)
        elif file_extension == '.txt':
            text = self.extract_text_from_txt(file_path)
        else:
            logger.warning(f"Unsupported file type: {file_extension}")
            return []
        
        if not text.strip():
            logger.warning(f"No text extracted from {filename}")
            return []
        
        # Split text into chunks
        chunks = self.text_splitter.split_text(text)
        
        # Create documents with metadata
        documents = []
        for i, chunk in enumerate(chunks):
            documents.append({
                'content': chunk,
                'metadata': {
                    'filename': filename,
                    'file_type': file_extension,
                    'chunk_index': i,
                    'total_chunks': len(chunks)
                }
            })
        
        return documents